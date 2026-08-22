"""GAIL PE cross-reference master — the grade-equivalence map.

This is the join every comparison depends on: it says which competitor grade is
the like-for-like substitute for a GAIL grade. Without it the engine can compare
prices but not products.

Two caveats travel with the data and are carried through to the output rather
than being silently dropped:

  * It was last updated 31-Dec-2023 and its competitor grades were read off
    price lists from Dec 2023 / Dec 2025 / Jun 2026 — all older than the Aug 2026
    circulars this engine prices from. A grade that has since been renamed or
    withdrawn will map to something that no longer exists.
  * Each row carries its own confidence flag (H/M/L). Medium rows are ones where
    an HMEL code changed between 2023 and 2025.

The workbook also lists 14 application sectors where GAIL has no grade at all;
those are the gaps a sales officer cannot quote against, so they are extracted
too rather than left in a spreadsheet nobody opens.
"""

from __future__ import annotations

import openpyxl

# Competitor columns, in workbook order, on both cross-reference sheets.
PRODUCERS = ["BCPL", "RIL", "HPL", "IOCL", "OPaL", "HMEL"]
DASH = {"—", "-", "–", ""}
# Multi-grade cells separate alternatives with a middle dot or a slash.
SEPARATORS = ["·", "/", "•"]


def _split_grades(cell) -> list[str]:
    if cell is None:
        return []
    text = str(cell).strip()
    if text in DASH:
        return []
    for sep in SEPARATORS[1:]:
        text = text.replace(sep, SEPARATORS[0])
    out = []
    for part in text.split(SEPARATORS[0]):
        part = part.strip().rstrip("*").strip()
        if part and part not in DASH:
            out.append(part)
    return out


def _sheet_rows(ws) -> list[dict]:
    """Rows of a cross-reference sheet, carrying the section heading above them."""
    out: list[dict] = []
    section = ""
    for row in ws.iter_rows(min_row=3, values_only=True):
        cells = list(row) + [None] * (16 - len(row))
        serial, grade = cells[0], cells[1]
        # Section headings occupy column A alone (INJECTION MOULDING, PIPE, ...).
        if serial and not grade:
            section = str(serial).strip()
            continue
        # Serials arrive as text, not numbers — the workbook stores the whole
        # sheet as strings, so an isinstance(int) test drops every row.
        if not grade or not str(serial or "").strip().isdigit():
            continue
        equivalents = {
            producer: _split_grades(cells[7 + i])
            for i, producer in enumerate(PRODUCERS)
        }
        out.append(
            {
                "gail_grade": str(grade).strip(),
                "section": section,
                "process": str(cells[2] or "").strip(),
                "application": str(cells[3] or "").strip(),
                "characteristic": str(cells[4] or "").strip(),
                "mfi": str(cells[5] or "").strip(),
                "density": str(cells[6] or "").strip(),
                "equivalents": equivalents,
                "international": _split_grades(cells[13]),
                "confidence": str(cells[14] or "").strip(),
                "source": str(cells[15] or "").strip(),
            }
        )
    return out


def load(path: str) -> dict:
    """{"hdpe": [...], "lldpe": [...], "gaps": [...], "provenance": {...}}"""
    book = openpyxl.load_workbook(path, data_only=True)

    gaps = []
    ws = book["Portfolio Gaps"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        gaps.append(
            {
                "polymer": str(row[0]).strip(),
                "sector": str(row[1] or "").strip(),
                "characteristic": str(row[2] or "").strip(),
                "ril": _split_grades(row[3]),
                "others": str(row[4] or "").strip(),
                "international": _split_grades(row[5]),
            }
        )

    provenance = {}
    ws = book["Summary Statistics"]
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row and row[0] and row[1] is not None:
            provenance[str(row[0]).strip()] = str(row[1]).strip()

    return {
        "hdpe": _sheet_rows(book["HDPE Cross-Reference"]),
        "lldpe": _sheet_rows(book["LLDPE Cross-Reference"]),
        "gaps": gaps,
        "provenance": provenance,
    }


def index_by_gail_grade(data: dict) -> dict[str, dict]:
    """Flatten both sheets into one lookup keyed by GAIL grade."""
    out: dict[str, dict] = {}
    for polymer in ("hdpe", "lldpe"):
        for row in data[polymer]:
            entry = dict(row)
            entry["polymer"] = polymer.upper()
            out[entry["gail_grade"]] = entry
    return out
